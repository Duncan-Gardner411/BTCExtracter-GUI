
# Written By Duncan Gardner, University of New Mexico,
# working under Dr.Ricardo Gonzales-Pinzon, 
# Contact me at dugardner@unm.edu or duncangardner411@gmail.com




from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
    )
import os
from tkinter import *
from tkinter import messagebox
from openpyxl.reader.excel import load_workbook


#Defines what file is used as the template
TemplateName = "DataBaseFormat.xlsx"
FormatWorkbook = load_workbook(filename = TemplateName)
FormattedSheet = FormatWorkbook["Sheet1"]


#List of rows in the format that need to accept float values
floatrows = [6,7,8,9,10,11,13,14,15,16,17,18,20,21,22,23,24,26,27,28,29,31]
UnitConvRows = [6,7,8,9,10,11,13,14,15,16,18,20,22,23,24,26,27,29,31]
#List of rows in the format that need to accept string values
stringrows = [1,3,4,5,12,19,25,30,32,33,39,40]

#List of Possible Units for use in program
VelocityUnits = ["ft/s", "mph","kph", "m/s", "m/min", "cm/s", ""]
VelocityRows = [11,23,24]
DistanceUnits = ["km", "m", "cm", "mm", "ft", "in", "mi", "yd", ""]
DistanceRows = [6,7,9,20]
AreaUnits = ["m^2", "cm^2", "ft^2", "in^2", "yd^2", ""]
AreaRows = [8]
FlowRateUnits = ["cms", "Lps", "Lpm", "cfs", "gps", "gpm", ""]
FlowRateRows = [10,22]
SlopeUnits = ["%", "ft/1000ft", ""]
SlopeRows =[18]
MassUnits = ["g", "kg", "mg", "lb", ""]
MassRows = [15]
TemperatureUnits = ["C", "F", ""]
TemperatureRows =[26]
TimeUnits = ["seconds", "minutes", "hours", ""]
TimeRows =[31]
ConcentrationUnits = ["%","ug/L", "mg/L", "g/L", "ppb", "ppm",""]
ConcentrationRows = [14,16,27,29]
PercentUnits = ["%", ""]
PercentRows = [13]

#Unit Conversion Dictionary
Conv = {
    #Distance
    'ftTOft' : 1,
    'kmTOft' : 3280.84,
    'mTOft' : 3.281,
    'cmTOft' : 0.03281,
    'mmTOft' : 0.003281,
    'inTOft' : 0.08333,
    'ydTOdt' : 3,
    'miTOft' : 5280,
    #Velocity
    'ft/sTOft/s' : 1,
    'm/sTOft/s' : 3.281,
    'mphTOft/s' : 1.4667,
    'kphTOft/s' : 0.9113,
    'm/minTOft/s' : 0.05468,
    'cm/sTOft/s' : 0.03281,
    #Area
    'ft^2TOft^2' : 1,
    'm^2TOft^2' : 10.7639,
    'cm^2TOft^2' : 0.001076,
    'in^2TOft^2' : 0.006944,
    'yd^2TOft^2' : 9,
    #FlowRate
    'cfsTOcfs' : 1,
    'cmsTOcfs' : 35.3147,
    'LpsTOcfs' : 0.03531,
    'LpmTOcfs' : 0.0005886,
    'gpsTOcfs' : 0.1337,
    'gpmTOcfs' : 0.002228,
    #Slope
    'ft/1000ftTOft/1000ft' : 1,
    '%TOft/1000ft' : 10,
    'TOft/1000ft' : 1000,
    #Mass
    'kgTOkg' : 1,
    'gTOkg' : .001,
    'mgTOkg' : .000001,
    'lbTOkg' : 0.4536,
    #Temperature
    'FTOF' : 1,
    'CTOF' : 1.8, #ALSO APPLIES ADDITION OF 32 IN LINE
    #Time
    'hoursTOhours' : 1,
    'secondsTOhours' : 0.00027778,
    'minutesTOhours' : 0.016667,
    #Concentration
    'ppbTOppb' : 1,
    '%TOppb' : 10000000,
    'TOppb' :  100000,
    'ppmTOppb' : 1000,
    'g/LTOppb' : 1000000,
    'mg/LTOppb' : 1000,
    'ug/LTOppb' : 1,
    #Percent
    '%TO%' : 1,
    'TO%' : 100
    }

#Dictionary relating row number to the type of value being placed there
RowDictionary = {
    1 : "Date of Experiment Ex:12/31/2022",
    3 : "River Surveyed",
    4 : "Latitude" + " Ex:46\N{DEGREE SIGN} 55' N",
    5 : "Longitude" + " Ex:110\N{DEGREE SIGN} 52' W",
    6 : "Channel Width",
    7 : "Channel Depth",
    8 : "Channel Cross Sectional Area",
    9 : "Distance from Tracer Injection",
    10 : "Flow Rate",
    11: "Flow Velocity",
    12: "Type of Tracer",
    13: "Tracer Recovery %",
    14: "Concentration of Tracer Injection",
    15: "Mass of Tracer Injection",
    16: "Background Tracer Concentration",
    17: "River Order",
    18: "Bed Slope",
    19: "Bed Material",
    20: "Thickness of Bed Material",
    21: "Manning's n",
    22: "Source Location Flow Rate",
    23: "Source Location Velocity",
    24: "Shear Velocity",
    25: "Vegetation Notes",
    26: "Temperature",
    27: "Total Dissolved Solids",
    28: "pH",
    29: "Concentration Level of Detection",
    30: "BVP or IUP?",
    31: "Duration of Tracer Injection",
    32: "Tracer Mixing Notes @ Injection Site",
    33: "Monitoring Method",
    39: "Primary Reference",
    40: "Additional Notes",
    }
#Form a list containing all unit lists for use in frop down
allunits = [DistanceUnits, 
            AreaUnits,
            VelocityUnits,
            FlowRateUnits, 
            SlopeUnits, 
            MassUnits, 
            TemperatureUnits, 
            TimeUnits, 
            ConcentrationUnits,
            PercentUnits,
            ]

#Function that saves the values
def SaveValue():
    global CurrentRow
    global SelectedUnit
    EnteredValue = ValueEntry.get()
    if CurrentRow in floatrows and EnteredValue != "":
        EnteredValue = float(EnteredValue)
    FormattedSheet.cell(CurrentRow, 3).value = SelectedUnit.get()
    FormattedSheet.cell(CurrentRow, 4).value = EnteredValue


CurrentRow = 1
CurrentlyAsked = RowDictionary[CurrentRow]
#function that advances to next row, skipping over some and loopiong back to the bottom if it goes above 39
def RowAdvance():
    global CurrentRow
    global CurrentlyAsked
    if CurrentRow in floatrows and ValueEntry.get() != "":
        tempcheck = ValueEntry.get()
        try:
           float(tempcheck)
        except:
            messagebox.showerror('Python Error', 'Error, Please enter a float value')
            return

        
    if CurrentRow not in [1, 33, 40]:
        CurrentRow += 1
    elif CurrentRow == 33:
        CurrentRow = 39
    elif CurrentRow == 1:
        CurrentRow = 3
    else:
        CurrentRow = 1
    CurrentlyAsked = RowDictionary[CurrentRow]
#function that return to previous row, skipping over some and looping back to the top if it goes below 1
def RowPrevious():
    global CurrentRow
    global CurrentlyAsked
    if CurrentRow in floatrows and ValueEntry.get() != "":
        tempcheck = ValueEntry.get()
        try:
           float(tempcheck)
        except:
            messagebox.showerror('Python Error', 'Error, Please enter a float value')
            return
    if CurrentRow not in [1, 3, 39]:
        CurrentRow -= 1
    elif CurrentRow == 39:
        CurrentRow = 33
    elif CurrentRow == 3:
        CurrentRow = 1
    else:
        CurrentRow = 40
    CurrentlyAsked = RowDictionary[CurrentRow]

#Function that changes the unit drop down to show propper units
def ChangeUnitDropDown():
    global CurrentPossibleUnits
    global CurrentRow
    global SelectedUnit
    if CurrentRow in VelocityRows: CurrentPossibleUnits = VelocityUnits
    elif CurrentRow in DistanceRows: CurrentPossibleUnits = DistanceUnits
    elif CurrentRow in AreaRows: CurrentPossibleUnits = AreaUnits
    elif CurrentRow in FlowRateRows: CurrentPossibleUnits = FlowRateUnits
    elif CurrentRow in SlopeRows: CurrentPossibleUnits = SlopeUnits
    elif CurrentRow in MassRows: CurrentPossibleUnits = MassUnits
    elif CurrentRow in TemperatureRows: CurrentPossibleUnits = TemperatureUnits
    elif CurrentRow in ConcentrationRows: CurrentPossibleUnits = ConcentrationUnits
    elif CurrentRow in TimeRows: CurrentPossibleUnits = TimeUnits
    elif CurrentRow in PercentRows: CurrentPossibleUnits = PercentUnits
    else: CurrentPossibleUnits = []
    menu = UnitDrop['menu']
    menu.delete(0, 'end')
    #After clearing the dropdown, adds the possible units to the new drop down.
    for name in CurrentPossibleUnits:
        # I found this on stackoverflow, I have no idea why the arguments have repeated use of 'name' 
        menu.add_command(label=name, command=lambda name=name: SelectedUnit.set(name))
    UnitNeeded = FormattedSheet.cell(CurrentRow, 3).value
    if UnitNeeded == None:
       UnitNeeded = ""
    SelectedUnit.set(UnitNeeded)


#Defines the list of actions taken when buttons are pressed
def AdvanceButtonClicked():
    global CurrentRow
    SaveValue()
    RowAdvance()
    ChangeUnitDropDown()
    CurrentAskedLabel.config(text = CurrentlyAsked)
    ValueEntry.delete(0, END)
    number = FormattedSheet.cell(CurrentRow,4).value
    if number == None: number = ""
    number = str(number)
    ValueEntry.insert(0, number)


def PreviousButtonClicked():
    global CurrentRow
    SaveValue()
    RowPrevious()
    ChangeUnitDropDown()
    CurrentAskedLabel.config(text = CurrentlyAsked)
    ValueEntry.delete(0, END)
    number = FormattedSheet.cell(CurrentRow,4).value
    if number == None: number = ""
    number = str(number)
    ValueEntry.insert(0, number)


#Change the widths of excel columns to match format.
def SetColumnWidths():
    FormattedSheet.column_dimensions["A"].width = 4.3
    FormattedSheet.column_dimensions["B"].width = 58
    FormattedSheet.column_dimensions["C"].width = 23
    FormattedSheet.column_dimensions["D"].width = 25.5
    FormattedSheet.column_dimensions["E"].width = 31


#Add Data to the Formatted Sheet by splitting the entered number along new lines and adding to the formatted sheet
def AddData():
    global DataLength
    FormattedSheet['B77'] = TimeUnit.get()
    FormattedSheet['C77'] = ConcUnit.get()
    FormattedSheet['D77'] = ConcUnit.get()
    FormattedSheet['E77'] = ConcUnit.get()
    DataString = ConcentrationEntry.get(1.0, 'end-1c')
    DataList = DataString.split('\n')
    row = 78 
    DataList.pop()
    DataLength =len(DataList)
    for val in DataList:
        if val == "": val = 0
        FormattedSheet.cell(row, 3).value = float(val)
        FormattedSheet.cell(row, 4).value = float(val)
        FormattedSheet.cell(row, 5).value = float(val)
        row += 1
    DataString = TimeEntry.get(1.0, 'end-1c')
    DataList = DataString.split('\n')
    row = 78
    DataList.pop()
    for val in DataList:
        if val == "": val = 0
        FormattedSheet.cell(row, 2).value = float(val)
        row += 1
    row = 78 + DataLength
    Repeat = True
    while Repeat:
        if FormattedSheet.cell(row, 2).value != None:
            FormattedSheet.cell(row, 2).value = ""
            FormattedSheet.cell(row, 3).value = ""
            FormattedSheet.cell(row, 4).value = ""
            FormattedSheet.cell(row, 5).value = ""
            row += 1
        else: Repeat = False

#Adds Chart to the converted sheet
def AddChart():
    chart = ScatterChart()
    chart.title = "BTC"
    chart.style = 2
    chart.x_axis.title = 'Time Since Injection (Hours)'
    chart.y_axis.title = 'Conservative Concentration (PPB)'
    LastRow = 77 + DataLength
    xvalues = Reference(TemporarySheet, min_col = 2, min_row = 78, max_row = LastRow)
    yvalues = Reference(TemporarySheet, min_col = 3, min_row = 78, max_row = LastRow)
    series = Series(yvalues, xvalues)
    chart.series.append(series)
    TemporarySheet.add_chart(chart, "F4")


def AddCitation():
    for row in range (3 ,41):
        if FormattedSheet.cell(row , 4).value not in [None, ""]:
            FormattedSheet.cell(row , 5).value = FormattedSheet.cell(39,4).value


#Converts to the final units using conversions dictionary
def ConvertToFinal():
    for Row in UnitConvRows:

        #Sets Final Unit Based on Current Row
        if Row in DistanceRows: FinalUnit = "ft"
        elif Row in VelocityRows: FinalUnit = "ft/s"
        elif Row in AreaRows: FinalUnit = "ft^2"
        elif Row in FlowRateRows: FinalUnit = "cfs"
        elif Row in SlopeRows: FinalUnit = "ft/1000ft"
        elif Row in MassRows: FinalUnit = "kg"
        elif Row in TemperatureRows: FinalUnit = "F"
        elif Row in TimeRows: FinalUnit = "hours"
        elif Row in ConcentrationRows: FinalUnit = "ppb"
        elif Row in PercentRows: FinalUnit = "%"

        if TemporarySheet.cell(Row, 4).value not in [None, ""] :
            UnitKey = TemporarySheet.cell(Row, 3).value
            if UnitKey == None: UnitKey = ""
            UnitKey = UnitKey + "TO" + FinalUnit
            print(TemporarySheet.cell(Row, 4).value)
            TemporarySheet.cell(Row , 4).value = round(TemporarySheet.cell(Row , 4).value * Conv[UnitKey] , 3)

            TemporarySheet.cell(Row , 3).value = FinalUnit
            #Exception for row needing mi
            if Row == 9:
                TemporarySheet.cell(Row , 4).value = round(TemporarySheet.cell(Row , 4).value / 5280 , 3)
                TemporarySheet.cell(Row , 3).value = "mi"
            
            #Expection for Temperature row
            if Row == 26:
                TemporarySheet.cell(Row, 4).value = TemporarySheet.cell(Row, 4).value + 32
    #Convert The Data
    StartRow = 78
    #Sets Conversion Factor For Time Data
    FinalUnit = "hours"
    UnitKey = TemporarySheet.cell(77, 2).value
    if UnitKey == None: UnitKey = ""
    UnitKey = UnitKey + "TO" + FinalUnit
    TimeConvFactor = Conv[UnitKey]
    TemporarySheet["B77"] = FinalUnit

    #Sets Conversion Factor For Concentration Data
    FinalUnit = "ppb"
    UnitKey = TemporarySheet.cell(77, 3).value
    if UnitKey == None: UnitKey = ""
    UnitKey = UnitKey + "TO" + FinalUnit
    ConcConvFactor = Conv[UnitKey]
    TemporarySheet["C77"] = FinalUnit
    TemporarySheet["D77"] = FinalUnit
    TemporarySheet["E77"] = FinalUnit

    #Convertes All Data
    for i in range(0 , DataLength):
        ROW = StartRow + i
        TemporarySheet.cell(ROW, 2).value = round(TemporarySheet.cell(ROW, 2).value * TimeConvFactor , 3)
        TemporarySheet.cell(ROW, 3).value = round(TemporarySheet.cell(ROW, 3).value * ConcConvFactor , 3)
        TemporarySheet.cell(ROW, 4).value = round(TemporarySheet.cell(ROW, 4).value * ConcConvFactor , 3)
        TemporarySheet.cell(ROW, 5).value = round(TemporarySheet.cell(ROW, 5).value * ConcConvFactor , 3)


#Runs final convertions and saves when button is clicked, either closing window or remaining open
def SaveAndQuitClicked():
    SaveButtonClicked()
    raise SystemExit

def SaveButtonClicked():
    global TemporarySheet
    SaveValue()
    #Creates the name of the excel sheet by getting the first 3 letters of river and an abbreviated date, checks if format is good
    First3OfRiver = 'XXX'
    if FormattedSheet.cell(3, 4).value != None:
        First3OfRiver = FormattedSheet.cell(3, 4).value[0 : 3]

    First3OfRiver = First3OfRiver.upper()
    DateSplit = FormattedSheet.cell(1, 4).value
    DateSplit = DateSplit.split("/")
    SectionNumber = TestNumberEntry.get()
    try:
        print(DateSplit)
        Month = DateSplit[0]
        Day = DateSplit[1]
        Year = DateSplit[2]
        if Month == "": Month = 'xx'
        if Day == "": Day = 'xx'
        if Year == "": Year = 'xx'
        print(Year)
        Year = Year[2: 4]
        print(Year)
        FormattedName = "%s %s-%s-%s S%s NonConverted.xlsx" % (First3OfRiver, Month, Day, Year, SectionNumber)
        ConvertedName = "%s %s-%s-%s S%s.xlsx" % (First3OfRiver, Month, Day, Year, SectionNumber)
        SheetTitle = "%s %s-%s-%s S%s" % (First3OfRiver, Month, Day, Year, SectionNumber)
    except:
        messagebox.showerror('Python Error', "Date Not Present or Correctly Formatted")
        return
    Year = Year[2: 4]
    if CurrentRow in floatrows and ValueEntry.get() != "":
        tempcheck = ValueEntry.get()
        try:
           float(tempcheck)
        except:
            messagebox.showerror('Python Error', 'Error, Please enter a float value')
            return

    AddData()
    AddCitation()
    SetColumnWidths()
    MonthCell = FormattedSheet["D1"]
    FormattedSheet["D1"] = ""
    FormattedSheet.title = SheetTitle
    FormatWorkbook.save(filename = FormattedName)

    TemporaryWB = load_workbook(filename = FormattedName)
    TemporarySheet = TemporaryWB[SheetTitle]


    ConvertToFinal()
    AddChart()
    TemporaryWB.save(filename = ConvertedName)

    #Resets the temp cell after saving
    #FormattedSheet["D1"] = MonthCell



#initialize the UI window
root = Tk()
#Label the window, size it and add a frame
root.title('UNM BTC Data Entry')
root.geometry("1050x650")
frame = LabelFrame(root, text = "Please Enter Values as Prompted", font = ('Aerial', 15))
frame.place(in_=root, relx =.01, rely =.01, width = 1000, height = 600)

CurrentAskedLabel = Label(frame, text = CurrentlyAsked, font = ('Times New Roman', 14), padx = 5, wraplength = 200, justify = "left")
CurrentAskedLabel.place(height = 100, width = 200, relx = .01, rely = 0.03)

#Creats labels, entry boxes and buttons for data entry

#initialize unit variable and set drop down variable type
CurrentPossibleUnits = []
SelectedUnit = StringVar()

#Creates Labels for drop down and value entry
ValueEntryLabel = Label(frame, text = "Enter Value Below", font = ('Times New Roman', 14), padx = 20)
UnitEntryLabel = Label(frame, text = "Select Unit From Drop Down", font = ('Times New Roman', 14), padx = 20)
BTCDataEntryLabel = Label(frame,text = "Copy Paste From Excel Directly Into The Boxes Below", font = ('Times New Roman', 14), padx = 20)
ConcentrationEntryLabel = Label(frame, text = "Concentration", font = ('Times New Roman', 14), padx = 20)
TimeEntryLabel = Label(frame, text = "Time After Injection", font = ('Times New Roman', 14), padx = 20)
TestNumberLabel =Label(frame, text = "Of Tests on This River, What Number is This", font = ('Times New Roman', 14), padx = 20, wraplength = 200, justify = "left")

TestNumberLabel.place(height = 50, width = 250, relx = .63, rely =0.47 )
BTCDataEntryLabel.place(height = 50, width = 450, relx = .15, rely = 0.4)
ConcentrationEntryLabel.place(height = 50, width = 200, relx = .15, rely = .47)
TimeEntryLabel.place(height = 50, width = 250, relx = .37, rely = .47)

ValueEntryLabel.place(height = 50, width = 200, relx = .25, rely = 0)
UnitEntryLabel.place(height = 50, width = 250, relx = .50, rely = 0)

#Creates Drop Down for Unit Selection
UnitDrop = OptionMenu(frame, SelectedUnit, [])
UnitDrop.place(height = 25, width = 150, relx = .55, rely = .1)

#Creates Entry Box for values
ValueEntry = Entry(frame, width = 200, font = ('Times New Roman', 14), bg = "lightgrey")
ValueEntry.place(height = 20, width = 200, relx = .25, rely = .1)

#Creates button to advance row
AdvanceButton = Button(frame, text = "Advance", font = ('Times New Roman', 14), command = AdvanceButtonClicked)
AdvanceButton.place(height = 50, width = 150, relx = .55, rely = .2)

#Creates button to retreat a row
PreviousButton = Button(frame, text = "Previous", font = ('Times New Roman', 14), command = PreviousButtonClicked)
PreviousButton.place(height = 50, width = 150, relx = .25, rely = .2)

#Concentration Entry Box
ConcentrationEntry = Text(frame, width =200, font = ('Times New Roman', 14), bg = "lightgrey")
ConcentrationEntry.place(height = 70, width = 200, relx =0.15, rely= .55 )

#Time after Injection Entry Box
TimeEntry = Text(frame, width =200, font = ('Times New Roman', 14), bg = "lightgrey")
TimeEntry.place(height = 70, width = 200, relx =0.4, rely= .55 )

#Test Number Entry Box
TestNumberEntry = Entry(frame, width = 50,  font = ('Times New Roman', 14), bg = "lightgrey")
TestNumberEntry.place(height = 30, width = 50, relx =0.72, rely= .55 )

#Create Save And Quit Button
SaveButton = Button(frame, text = "Save and Quit", font = ('Times New Roman', 14), command = SaveAndQuitClicked)
SaveButton.place(height = 50, width = 150, relx = .55, rely = .85)

#Create Save And Quit Button
SaveButton = Button(frame, text = "Save", font = ('Times New Roman', 14), command = SaveButtonClicked)
SaveButton.place(height = 50, width = 150, relx = .25, rely = .85)

#Unit Drop Down for Concentration
ConcUnit = StringVar()
ConcDrop = OptionMenu(frame, ConcUnit, *ConcentrationUnits)
ConcDrop.place(height = 25, width = 200, relx = .15, rely = .65)

#Unit Drop Down for Time 
TimeUnit = StringVar()
TimeDrop = OptionMenu(frame, TimeUnit, *TimeUnits)
TimeDrop.place(height = 25, width = 200, relx = .4, rely = .65)


#Binds Enter Key to Advancing the row
root.bind('<Return>',lambda event:AdvanceButtonClicked())



#loops over the UI reacting to inputs
ChangeUnitDropDown()
root.mainloop()